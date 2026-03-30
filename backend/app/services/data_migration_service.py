"""Data Migration Service — Sage 100 → Bridgeable import pipeline.

Handles parsing and importing:
- Chart of Accounts (COA) CSV
- Customer list XLSX
- AR Aging CSV
- Vendor list XLSX
- AP Aging CSV
"""

import csv
import io
import re
import uuid
from datetime import date, datetime, timezone
from decimal import Decimal, InvalidOperation
from typing import Generator

import openpyxl

from app.models.accounting_analysis import TenantGLMapping
from app.models.customer import Customer
from app.models.data_migration import DataMigrationRun
from app.models.invoice import Invoice, InvoiceLine
from app.models.vendor import Vendor
from app.models.vendor_bill import VendorBill
from app.models.vendor_bill_line import VendorBillLine


# ---------------------------------------------------------------------------
# Category mappings
# ---------------------------------------------------------------------------

SAGE_CATEGORY_MAP: dict[str, str] = {
    "CURRENT ASSETS": "current_asset",
    "FIXED ASSETS": "fixed_asset",
    "CURRENT LIABILITIES": "current_liability",
    "LONG TERM LIABILITIES": "long_term_liability",
    "EQUITY": "equity",
    "INCOME": "revenue",
    "COST OF SALES": "cogs",
    "MANUFACTURING EXPENSE": "expense",
    "ADMINISTRATIVE & SELLING EXPENSE": "expense",
    "OTHER INCOME & EXPENSES": "other_income",
    "PROVISION FOR TAXES": "tax_expense",
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def parse_amount(s: str | None) -> Decimal:
    """Parse a Sage-formatted amount string into a Decimal.

    Handles:
    - Comma-separated numbers:  "25,462.31"  -> Decimal("25462.31")
    - Trailing-minus negatives: "1,134.80-"  -> Decimal("-1134.80")
    - Empty / None              ""           -> Decimal("0")
    """
    if not s:
        return Decimal("0")
    s = s.strip()
    if not s:
        return Decimal("0")
    try:
        negative = s.endswith("-")
        cleaned = s.rstrip("-").replace(",", "")
        if not cleaned:
            return Decimal("0")
        value = Decimal(cleaned)
        return -value if negative else value
    except (InvalidOperation, ValueError):
        return Decimal("0")


def parse_date(s: str | None) -> datetime | None:
    """Parse a Sage date string (M/D/YYYY) into a UTC-aware datetime."""
    if not s:
        return None
    s = s.strip()
    if not s:
        return None
    try:
        dt = datetime.strptime(s, "%m/%d/%Y")
        return dt.replace(tzinfo=timezone.utc)
    except (ValueError, OverflowError):
        return None


# ---------------------------------------------------------------------------
# FILE 1: COA CSV parser
# ---------------------------------------------------------------------------


def parse_sage_coa(file_content: bytes) -> list[dict]:
    """Parse a Sage 100 Chart of Accounts CSV export.

    CSV row layout (after csv.reader):
    [0]  "Chart of Accounts"
    [1]  company name
    [9]  sage_category  (e.g. "CURRENT ASSETS")
    [10] account_number (4–5 digit number as string)
    [11] description
    [12] status         ("Active" / "Inactive" / "Deleted")
    """
    text = file_content.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))

    accounts: list[dict] = []
    seen: set[str] = set()

    for row in reader:
        if len(row) < 13:
            continue
        if row[0].strip() != "Chart of Accounts":
            continue

        account_number = row[10].strip()
        if not account_number:
            continue
        try:
            int(account_number)
        except ValueError:
            continue

        if account_number in seen:
            continue
        seen.add(account_number)

        description = row[11].strip()
        status_raw = row[12].strip().lower()  # "active", "inactive", "deleted"
        sage_category = row[9].strip().upper()

        bridgeable_account_type = SAGE_CATEGORY_MAP.get(sage_category)
        if bridgeable_account_type is not None:
            confidence = 1.0
        else:
            # Try a fuzzy match — check if any key is contained in the category
            matched = None
            for key, val in SAGE_CATEGORY_MAP.items():
                if key in sage_category or sage_category in key:
                    matched = val
                    break
            if matched:
                bridgeable_account_type = matched
                confidence = 0.7
            else:
                bridgeable_account_type = "other"
                confidence = 0.0

        accounts.append(
            {
                "account_number": account_number,
                "description": description,
                "status": status_raw,
                "sage_category": row[9].strip(),
                "bridgeable_account_type": bridgeable_account_type,
                "confidence": confidence,
            }
        )

    return accounts


# ---------------------------------------------------------------------------
# FILE 2: Customer XLSX parser
# ---------------------------------------------------------------------------


def parse_sage_customers(file_content: bytes) -> list[dict]:
    """Parse a Sage 100 Customer List XLSX export.

    Row 0 (index): title header "Customer List" — skip
    Row 1 (index): column headers
    Row 2+ (index): data

    Customer No. format: "00-AAL001 " → strip → split("-", 1) → ["00", "AAL001"]
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []

    # Row index 1 = headers
    header_row = rows[1]
    header_map: dict[str, int] = {}
    for idx, cell in enumerate(header_row):
        if cell is not None:
            header_map[str(cell).strip()] = idx

    def get_cell(row: tuple, header: str) -> str:
        idx = header_map.get(header)
        if idx is None or idx >= len(row):
            return ""
        val = row[idx]
        return str(val).strip() if val is not None else ""

    customers: list[dict] = []

    # Data starts at row index 2
    for row in rows[2:]:
        customer_no_raw = get_cell(row, "Customer No.")
        if not customer_no_raw:
            continue

        customer_no = customer_no_raw.strip()
        parts = customer_no.split("-", 1)
        if len(parts) == 2:
            division = parts[0].strip()
            sage_id = parts[1].strip()
        else:
            division = "00"
            sage_id = customer_no.strip()

        name = get_cell(row, "Name")
        if not name:
            continue

        status = get_cell(row, "Status") or "Active"
        zip_code = get_cell(row, "ZIP/Postal Code")
        phone = get_cell(row, "Phone Number")
        customer_type_raw = get_cell(row, "CustomerType")
        price_level = get_cell(row, "PriceLevel")
        statement_cycle = get_cell(row, "StatementCycle")
        date_established = get_cell(row, "DateEstablished")

        # Determine customer_type
        if customer_type_raw == "PC":
            customer_type = "contractor"
        elif division == "10" or division == "11":
            customer_type = "funeral_home"
        elif division == "00":
            # If PC type, already caught above. Otherwise default for div 00:
            customer_type = "funeral_home"
        else:
            customer_type = "funeral_home"

        # CustomerType field overrides division logic for contractors
        if customer_type_raw == "PC":
            customer_type = "contractor"

        customers.append(
            {
                "sage_customer_no": customer_no,
                "name": name,
                "status": status,
                "zip": zip_code,
                "phone": phone,
                "customer_type": customer_type,
                "price_level": price_level,
                "statement_cycle": statement_cycle,
                "date_established": date_established,
                "division": division,
                "sage_id": sage_id,
            }
        )

    return customers


# ---------------------------------------------------------------------------
# FILE 3: Vendor XLSX parser
# ---------------------------------------------------------------------------


def parse_sage_vendors(file_content: bytes) -> list[dict]:
    """Parse a Sage 100 Vendor List XLSX export.

    Row 0 (index): title header "Vendor List" — skip
    Row 1 (index): column headers
    Row 2+ (index): data
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []

    header_row = rows[1]
    header_map: dict[str, int] = {}
    for idx, cell in enumerate(header_row):
        if cell is not None:
            header_map[str(cell).strip()] = idx

    def get_cell(row: tuple, header: str) -> str:
        idx = header_map.get(header)
        if idx is None or idx >= len(row):
            return ""
        val = row[idx]
        return str(val).strip() if val is not None else ""

    vendors: list[dict] = []

    for row in rows[2:]:
        vendor_no = get_cell(row, "Vendor No.")
        if not vendor_no:
            continue

        name = get_cell(row, "Name")
        if not name:
            continue

        status = get_cell(row, "Status") or "Active"
        zip_code = get_cell(row, "ZIP/Postal Code")
        phone = get_cell(row, "Phone Number")

        vendors.append(
            {
                "sage_vendor_no": vendor_no.strip(),
                "name": name,
                "status": status,
                "zip": zip_code,
                "phone": phone,
            }
        )

    return vendors


# ---------------------------------------------------------------------------
# FILE 4: AR Aging CSV parser
# ---------------------------------------------------------------------------


def parse_sage_ar_aging(file_content: bytes) -> tuple[list[dict], list[dict]]:
    """Parse a Sage 100 AR Aged Invoice Report CSV.

    Returns (invoices, customers) tuple.

    Row layout (after csv.reader):
    [0]  "Accounts Receivable Aged Invoice Report"
    [4]  "Division Number:  00 PRECAST DIVISION"
    [21] customer_no
    [22] customer_name
    [23] "Contact:"
    [24] contact_name
    [25] "Phone:"
    [26] phone
    [29] "Credit Limit:"
    [30] credit_limit
    [32] invoice_date
    [33] invoice_number
    [34] due_date
    [37] discount_amount
    [38] balance (may be "1,134.80-" for negative)
    [39] current_amount
    [40] one_month
    [41] two_months
    [42] three_months
    [43] four_months
    [44] days_delinquent
    """
    text = file_content.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))

    invoices: list[dict] = []
    customers_map: dict[str, dict] = {}  # keyed by customer_no

    for row in reader:
        if len(row) < 45:
            continue
        if row[0].strip() != "Accounts Receivable Aged Invoice Report":
            continue

        # Extract division from field [4]
        division_match = re.search(r"Division Number:\s+(\d+)", row[4])
        division = division_match.group(1) if division_match else ""

        customer_no = row[21].strip()
        if not customer_no:
            continue
        # Skip totals / summary rows
        low = customer_no.lower()
        if "totals:" in low or "report" in low or "division" in low:
            continue

        customer_name = row[22].strip()
        contact_name = row[24].strip() if len(row) > 24 else ""
        phone = row[26].strip() if len(row) > 26 else ""

        # Credit limit
        credit_limit = Decimal("0")
        if len(row) > 30 and row[29].strip() == "Credit Limit:":
            credit_limit = parse_amount(row[30])

        invoice_date = parse_date(row[32]) if len(row) > 32 else None
        invoice_number = row[33].strip() if len(row) > 33 else ""
        due_date = parse_date(row[34]) if len(row) > 34 else None

        discount_amount = parse_amount(row[37]) if len(row) > 37 else Decimal("0")
        balance = parse_amount(row[38]) if len(row) > 38 else Decimal("0")
        current_amount = parse_amount(row[39]) if len(row) > 39 else Decimal("0")
        one_month = parse_amount(row[40]) if len(row) > 40 else Decimal("0")
        two_months = parse_amount(row[41]) if len(row) > 41 else Decimal("0")
        three_months = parse_amount(row[42]) if len(row) > 42 else Decimal("0")
        four_months = parse_amount(row[43]) if len(row) > 43 else Decimal("0")

        days_raw = row[44].strip() if len(row) > 44 else ""
        try:
            days_delinquent = int(days_raw.lstrip("-")) if days_raw.lstrip("-").isdigit() else 0
        except (ValueError, AttributeError):
            days_delinquent = 0

        # Check for credit hold / limit exceeded flags in remaining columns
        rest = " ".join(row[45:]) if len(row) > 45 else ""
        on_credit_hold = "*** On Credit Hold ***" in rest
        credit_limit_exceeded = "*** Credit Limit Exceeded ***" in rest

        # Skip empty invoice numbers or totals-like invoice numbers
        if not invoice_number:
            continue
        inv_low = invoice_number.lower()
        if "totals:" in inv_low or "division" in inv_low or invoice_number == "PP-PP":
            # PP-PP is a placeholder sometimes used for prepayments — keep if balance non-zero
            if invoice_number == "PP-PP" and balance == Decimal("0"):
                continue
            elif "totals:" in inv_low or "division" in inv_low:
                continue

        # Build / update customer record (first occurrence wins for basic fields)
        if customer_no not in customers_map:
            customers_map[customer_no] = {
                "sage_customer_no": customer_no,
                "customer_name": customer_name,
                "contact_name": contact_name,
                "phone": phone,
                "credit_limit": credit_limit,
                "on_credit_hold": on_credit_hold,
                "credit_limit_exceeded": credit_limit_exceeded,
                "division": division,
            }
        else:
            # Update flags if found on any row for this customer
            if on_credit_hold:
                customers_map[customer_no]["on_credit_hold"] = True
            if credit_limit_exceeded:
                customers_map[customer_no]["credit_limit_exceeded"] = True

        invoices.append(
            {
                "sage_customer_no": customer_no,
                "customer_name": customer_name,
                "invoice_number": invoice_number,
                "invoice_date": invoice_date,
                "due_date": due_date,
                "discount_amount": discount_amount,
                "balance": balance,
                "current_amount": current_amount,
                "one_month": one_month,
                "two_months": two_months,
                "three_months": three_months,
                "four_months": four_months,
                "days_delinquent": days_delinquent,
                "on_credit_hold": on_credit_hold,
                "credit_limit_exceeded": credit_limit_exceeded,
                "division": division,
            }
        )

    return invoices, list(customers_map.values())


# ---------------------------------------------------------------------------
# FILE 5: AP Aging CSV parser
# ---------------------------------------------------------------------------


def parse_sage_ap_aging(file_content: bytes) -> list[dict]:
    """Parse a Sage 100 AP Aged Invoice Report CSV.

    Row layout (after csv.reader):
    [0]  "Accounts Payable Aged Invoice Report"
    [21] "AGW001  SUBURBAN PROPANE"  (vendor_no + spaces + vendor_name combined)
    [22] empty
    [23] invoice_number
    [24] invoice_date
    [25] due_date
    [27] hold ("Yes"/"No")
    [28] invoice_balance
    [29] discount_amount
    [30] current_amount
    [31] one_month
    [32] two_months
    [33] three_months
    [34] four_months
    """
    text = file_content.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))

    bills: list[dict] = []

    for row in reader:
        if len(row) < 35:
            continue
        if row[0].strip() != "Accounts Payable Aged Invoice Report":
            continue

        vendor_field = row[21].strip()
        if not vendor_field:
            continue

        # Skip totals / summary rows
        low = vendor_field.lower()
        if "totals:" in low or "report" in low or "division" in low:
            continue

        # Split vendor_no and vendor_name on 2+ consecutive spaces
        parts = re.split(r"\s{2,}", vendor_field, 1)
        if len(parts) == 2:
            sage_vendor_no = parts[0].strip()
            vendor_name = parts[1].strip()
        else:
            sage_vendor_no = vendor_field.strip()
            vendor_name = ""

        if not sage_vendor_no:
            continue

        invoice_number = row[23].strip() if len(row) > 23 else ""
        if not invoice_number:
            continue
        inv_low = invoice_number.lower()
        if "totals:" in inv_low or "division" in inv_low:
            continue

        invoice_date = parse_date(row[24]) if len(row) > 24 else None
        due_date = parse_date(row[25]) if len(row) > 25 else None
        on_hold = row[27].strip().lower() == "yes" if len(row) > 27 else False
        invoice_balance = parse_amount(row[28]) if len(row) > 28 else Decimal("0")
        discount_amount = parse_amount(row[29]) if len(row) > 29 else Decimal("0")
        current_amount = parse_amount(row[30]) if len(row) > 30 else Decimal("0")
        one_month = parse_amount(row[31]) if len(row) > 31 else Decimal("0")
        two_months = parse_amount(row[32]) if len(row) > 32 else Decimal("0")
        three_months = parse_amount(row[33]) if len(row) > 33 else Decimal("0")
        four_months = parse_amount(row[34]) if len(row) > 34 else Decimal("0")

        bills.append(
            {
                "sage_vendor_no": sage_vendor_no,
                "vendor_name": vendor_name,
                "invoice_number": invoice_number,
                "invoice_date": invoice_date,
                "due_date": due_date,
                "on_hold": on_hold,
                "invoice_balance": invoice_balance,
                "discount_amount": discount_amount,
                "current_amount": current_amount,
                "one_month": one_month,
                "two_months": two_months,
                "three_months": three_months,
                "four_months": four_months,
            }
        )

    return bills


# ---------------------------------------------------------------------------
# Import: GL Accounts
# ---------------------------------------------------------------------------


def import_gl_accounts(
    db,
    tenant_id: str,
    parsed_accounts: list[dict],
    options: dict,
) -> dict:
    """Import parsed COA accounts into TenantGLMapping.

    options:
        include_inactive (bool, default False)
        overwrite_existing (bool, default False)
    """
    include_inactive: bool = options.get("include_inactive", False)
    overwrite_existing: bool = options.get("overwrite_existing", False)

    imported = 0
    skipped = 0
    errors: list[str] = []

    for acct in parsed_accounts:
        try:
            status = acct.get("status", "active")
            if not include_inactive and status == "inactive":
                skipped += 1
                continue
            if not include_inactive and status == "deleted":
                skipped += 1
                continue

            account_number = acct["account_number"]
            existing = (
                db.query(TenantGLMapping)
                .filter(
                    TenantGLMapping.tenant_id == tenant_id,
                    TenantGLMapping.account_number == account_number,
                )
                .first()
            )

            if existing is None:
                mapping = TenantGLMapping(
                    id=str(uuid.uuid4()),
                    tenant_id=tenant_id,
                    platform_category=acct["bridgeable_account_type"],
                    account_number=account_number,
                    account_name=acct["description"],
                    provider_account_id=account_number,
                    is_active=(status == "active"),
                )
                nested = db.begin_nested()
                try:
                    db.add(mapping)
                    db.flush()
                    nested.commit()
                    imported += 1
                except Exception as flush_err:
                    nested.rollback()
                    raise flush_err
            elif overwrite_existing:
                existing.platform_category = acct["bridgeable_account_type"]
                existing.account_name = acct["description"]
                existing.provider_account_id = account_number
                existing.is_active = status == "active"
                nested = db.begin_nested()
                try:
                    db.flush()
                    nested.commit()
                    imported += 1
                except Exception as flush_err:
                    nested.rollback()
                    raise flush_err
            else:
                skipped += 1

        except Exception as e:
            errors.append(f"Account {acct.get('account_number', '?')}: {str(e)}")

    try:
        db.commit()
    except Exception as e:
        db.rollback()
        errors.append(f"Commit error: {str(e)}")

    return {"imported": imported, "skipped": skipped, "errors": errors}


# ---------------------------------------------------------------------------
# Import: Customers
# ---------------------------------------------------------------------------


def import_customers(
    db,
    tenant_id: str,
    parsed_customers: list[dict],
    ar_customers: list[dict],
    options: dict,
) -> tuple[dict, dict]:
    """Import parsed customers into the Customer table.

    Returns (result_dict, customer_id_map) where
    customer_id_map = {sage_customer_no: bridgeable_customer_id}
    """
    include_inactive: bool = options.get("include_inactive", False)

    # Build AR data map: sage_customer_no -> ar_customer dict
    ar_map: dict[str, dict] = {c["sage_customer_no"]: c for c in ar_customers}

    imported = 0
    skipped = 0
    errors: list[str] = []
    customer_id_map: dict[str, str] = {}

    for cust in parsed_customers:
        try:
            status = cust.get("status", "Active")
            if not include_inactive and status == "Inactive":
                skipped += 1
                continue

            sage_customer_no = cust["sage_customer_no"]
            sage_id = cust.get("sage_id", sage_customer_no)  # part after "-"

            # Check if already imported by sage_customer_id
            existing = (
                db.query(Customer)
                .filter(
                    Customer.company_id == tenant_id,
                    Customer.sage_customer_id == sage_customer_no,
                )
                .first()
            )
            if existing is None:
                # Also check by account_number matching the sage_id part
                existing = (
                    db.query(Customer)
                    .filter(
                        Customer.company_id == tenant_id,
                        Customer.account_number == sage_id,
                    )
                    .first()
                )

            if existing is not None:
                customer_id_map[sage_customer_no] = existing.id
                skipped += 1
                continue

            ar_data = ar_map.get(sage_customer_no, {})
            customer_type = cust.get("customer_type", "funeral_home")
            division = cust.get("division", "00")

            billing_profile = (
                "monthly_statement" if customer_type == "funeral_home" else "invoice_on_order"
            )
            payment_terms = "net_30"
            account_status = "credit_hold" if ar_data.get("on_credit_hold") else "active"
            if status == "Inactive":
                account_status = "inactive"

            credit_limit = ar_data.get("credit_limit")
            if credit_limit is not None and credit_limit == Decimal("0"):
                credit_limit = None

            notes = f"Imported from Sage 100. Division: {division}"
            if customer_type:
                notes += f". Type: {customer_type}"

            customer = Customer(
                id=str(uuid.uuid4()),
                company_id=tenant_id,
                name=cust["name"],
                account_number=sage_id,
                phone=cust.get("phone") or None,
                zip_code=cust.get("zip") or None,
                contact_name=ar_data.get("contact_name") or None,
                credit_limit=credit_limit,
                account_status=account_status,
                billing_profile=billing_profile,
                payment_terms=payment_terms,
                customer_type=customer_type,
                sage_customer_id=sage_customer_no,
                notes=notes,
                is_active=(status != "Inactive"),
            )
            nested = db.begin_nested()
            try:
                db.add(customer)
                db.flush()
                nested.commit()
                customer_id_map[sage_customer_no] = customer.id
                imported += 1
            except Exception as flush_err:
                nested.rollback()
                raise flush_err

        except Exception as e:
            errors.append(f"Customer {cust.get('sage_customer_no', '?')}: {str(e)}")

    try:
        db.commit()
    except Exception as e:
        db.rollback()
        errors.append(f"Commit error: {str(e)}")

    return (
        {"imported": imported, "skipped": skipped, "errors": errors},
        customer_id_map,
    )


# ---------------------------------------------------------------------------
# Import: Open AR Invoices
# ---------------------------------------------------------------------------


def import_open_invoices(
    db,
    tenant_id: str,
    parsed_invoices: list[dict],
    customer_id_map: dict,
    cutover_date: date,
) -> dict:
    """Import open AR invoices as Invoice + InvoiceLine records."""
    imported = 0
    skipped = 0
    errors: list[str] = []
    warnings: list[str] = []

    cutover_note = f"Balance migrated from Sage 100 as of {cutover_date}"
    fallback_date = datetime.now(timezone.utc)

    for inv in parsed_invoices:
        try:
            sage_customer_no = inv["sage_customer_no"]
            invoice_number = inv["invoice_number"]

            customer_id = customer_id_map.get(sage_customer_no)
            if not customer_id:
                # AR aging CSV stores bare IDs (e.g. "AAL001") while customer XLSX
                # stores division-prefixed IDs (e.g. "00-AAL001").  Try prepending
                # the division extracted from this invoice row, then fall back to "00-".
                division = inv.get("division", "")
                if division:
                    customer_id = customer_id_map.get(f"{division}-{sage_customer_no}")
                if not customer_id:
                    customer_id = customer_id_map.get(f"00-{sage_customer_no}")
            if not customer_id:
                warnings.append(
                    f"Invoice {invoice_number}: customer {sage_customer_no} not found — skipped"
                )
                skipped += 1
                continue

            # Skip if already imported
            existing = (
                db.query(Invoice)
                .filter(
                    Invoice.company_id == tenant_id,
                    Invoice.sage_invoice_id == invoice_number,
                )
                .first()
            )
            if existing is not None:
                skipped += 1
                continue

            balance = inv["balance"]

            # Skip fully-paid invoices
            if balance == Decimal("0"):
                skipped += 1
                continue

            days_delinquent = inv.get("days_delinquent", 0)
            abs_balance = abs(balance)

            # Determine status — use "sent" (not "open") so the financials board
            # and AR aging queries find these invoices (they filter on sent/partial/overdue)
            if balance < 0:
                status = "sent"  # credit / overpayment — treat as sent with credit balance
            elif days_delinquent > 0:
                status = "overdue"
            else:
                status = "sent"

            number = f"SAGE-{invoice_number}"[:50]
            invoice_date = inv.get("invoice_date") or fallback_date
            due_date = inv.get("due_date") or fallback_date

            invoice = Invoice(
                id=str(uuid.uuid4()),
                company_id=tenant_id,
                number=number,
                customer_id=customer_id,
                status=status,
                invoice_date=invoice_date,
                due_date=due_date,
                subtotal=abs_balance,
                tax_rate=Decimal("0.0000"),
                tax_amount=Decimal("0"),
                total=abs_balance,
                amount_paid=Decimal("0"),
                notes=cutover_note,
                sage_invoice_id=invoice_number,
            )
            nested = db.begin_nested()
            try:
                db.add(invoice)
                db.flush()
                line = InvoiceLine(
                    id=str(uuid.uuid4()),
                    invoice_id=invoice.id,
                    description=cutover_note,
                    quantity=Decimal("1"),
                    unit_price=abs_balance,
                    line_total=abs_balance,
                    sort_order=0,
                )
                db.add(line)
                db.flush()
                nested.commit()
                imported += 1
            except Exception as flush_err:
                nested.rollback()
                raise flush_err

        except Exception as e:
            errors.append(f"Invoice {inv.get('invoice_number', '?')}: {str(e)}")

    try:
        db.commit()
    except Exception as e:
        db.rollback()
        errors.append(f"Commit error: {str(e)}")

    return {
        "imported": imported,
        "skipped": skipped,
        "errors": errors,
        "warnings": warnings,
    }


# ---------------------------------------------------------------------------
# Import: Vendors
# ---------------------------------------------------------------------------


def import_vendors(
    db,
    tenant_id: str,
    parsed_vendors: list[dict],
    options: dict,
) -> tuple[dict, dict]:
    """Import parsed vendors into the Vendor table.

    Returns (result_dict, vendor_id_map)
    """
    include_inactive: bool = options.get("include_inactive", False)

    imported = 0
    skipped = 0
    errors: list[str] = []
    vendor_id_map: dict[str, str] = {}

    for vend in parsed_vendors:
        try:
            status = vend.get("status", "Active")
            if not include_inactive and status != "Active":
                skipped += 1
                continue

            sage_vendor_no = vend["sage_vendor_no"]

            # Check if already imported
            existing = (
                db.query(Vendor)
                .filter(
                    Vendor.company_id == tenant_id,
                    Vendor.sage_vendor_id == sage_vendor_no,
                )
                .first()
            )
            if existing is not None:
                vendor_id_map[sage_vendor_no] = existing.id
                skipped += 1
                continue

            vendor = Vendor(
                id=str(uuid.uuid4()),
                company_id=tenant_id,
                name=vend["name"],
                account_number=sage_vendor_no,
                phone=vend.get("phone") or None,
                zip_code=vend.get("zip") or None,
                vendor_status="active" if status == "Active" else "inactive",
                sage_vendor_id=sage_vendor_no,
                is_active=(status == "Active"),
            )
            nested = db.begin_nested()
            try:
                db.add(vendor)
                db.flush()
                nested.commit()
                vendor_id_map[sage_vendor_no] = vendor.id
                imported += 1
            except Exception as flush_err:
                nested.rollback()
                raise flush_err

        except Exception as e:
            errors.append(f"Vendor {vend.get('sage_vendor_no', '?')}: {str(e)}")

    try:
        db.commit()
    except Exception as e:
        db.rollback()
        errors.append(f"Commit error: {str(e)}")

    return (
        {"imported": imported, "skipped": skipped, "errors": errors},
        vendor_id_map,
    )


# ---------------------------------------------------------------------------
# Import: Open AP Bills
# ---------------------------------------------------------------------------


def import_open_bills(
    db,
    tenant_id: str,
    parsed_bills: list[dict],
    vendor_id_map: dict,
    cutover_date: date,
) -> dict:
    """Import open AP bills as VendorBill + VendorBillLine records."""
    imported = 0
    skipped = 0
    errors: list[str] = []
    warnings: list[str] = []

    cutover_note = f"Balance migrated from Sage 100 as of {cutover_date}"
    fallback_date = datetime.now(timezone.utc)

    for bill_data in parsed_bills:
        try:
            sage_vendor_no = bill_data["sage_vendor_no"]
            invoice_number = bill_data["invoice_number"]

            vendor_id = vendor_id_map.get(sage_vendor_no)
            if not vendor_id:
                warnings.append(
                    f"Bill {invoice_number}: vendor {sage_vendor_no} not found — skipped"
                )
                skipped += 1
                continue

            # Skip if already imported (check by number pattern)
            bill_number = f"SAGE-{invoice_number}"[:50]
            existing = (
                db.query(VendorBill)
                .filter(
                    VendorBill.company_id == tenant_id,
                    VendorBill.number == bill_number,
                    VendorBill.vendor_id == vendor_id,
                )
                .first()
            )
            if existing is not None:
                skipped += 1
                continue

            balance = bill_data["invoice_balance"]
            if balance == Decimal("0"):
                skipped += 1
                continue

            abs_balance = abs(balance)
            invoice_date = bill_data.get("invoice_date") or fallback_date
            due_date = bill_data.get("due_date") or fallback_date

            bill = VendorBill(
                id=str(uuid.uuid4()),
                company_id=tenant_id,
                number=bill_number,
                vendor_id=vendor_id,
                vendor_invoice_number=invoice_number,
                status="pending",
                bill_date=invoice_date,
                due_date=due_date,
                subtotal=abs_balance,
                tax_amount=Decimal("0"),
                total=abs_balance,
                amount_paid=Decimal("0"),
                source="sage_migration",
                notes=cutover_note,
            )
            nested = db.begin_nested()
            try:
                db.add(bill)
                db.flush()
                line = VendorBillLine(
                    id=str(uuid.uuid4()),
                    bill_id=bill.id,
                    description=cutover_note,
                    quantity=Decimal("1"),
                    unit_cost=abs_balance,
                    amount=abs_balance,
                    sort_order=0,
                )
                db.add(line)
                db.flush()
                nested.commit()
                imported += 1
            except Exception as flush_err:
                nested.rollback()
                raise flush_err

        except Exception as e:
            errors.append(f"Bill {bill_data.get('invoice_number', '?')}: {str(e)}")

    try:
        db.commit()
    except Exception as e:
        db.rollback()
        errors.append(f"Commit error: {str(e)}")

    return {
        "imported": imported,
        "skipped": skipped,
        "errors": errors,
        "warnings": warnings,
    }


# ---------------------------------------------------------------------------
# Full migration orchestrator (generator / streaming)
# ---------------------------------------------------------------------------


def run_full_migration(
    db,
    tenant_id: str,
    files: dict,
    options: dict,
    cutover_date: date,
    initiated_by: str = "owner",
    extension_decisions: dict | None = None,
) -> Generator[dict, None, None]:
    """Run the full Sage 100 → Bridgeable migration pipeline.

    files = {
        "coa": bytes | None,
        "customers": bytes | None,
        "ar_aging": bytes | None,
        "vendors": bytes | None,
        "ap_aging": bytes | None,
    }

    Yields progress/status dicts for streaming to the client.
    """
    # Create the migration run record
    run = DataMigrationRun(
        id=str(uuid.uuid4()),
        tenant_id=tenant_id,
        status="in_progress",
        cutover_date=cutover_date,
        source_system="sage_100",
        initiated_by=initiated_by,
        started_at=datetime.now(timezone.utc),
    )
    db.add(run)
    db.commit()
    db.refresh(run)

    all_warnings: list[str] = []
    all_errors: list[str] = []

    # Build options subsets
    gl_options = {
        "include_inactive": options.get("include_inactive_accounts", False),
        "overwrite_existing": options.get("overwrite_existing", False),
    }
    customer_options = {
        "include_inactive": options.get("include_inactive_customers", False),
    }
    vendor_options = {
        "include_inactive": options.get("include_inactive_vendors", False),
    }

    customer_id_map: dict[str, str] = {}
    vendor_id_map: dict[str, str] = {}

    # ---- Step 1: GL Accounts ----
    if files.get("coa"):
        yield {"step": "gl_accounts", "status": "running", "progress": 0}
        try:
            parsed_coa = parse_sage_coa(files["coa"])
            yield {
                "step": "gl_accounts",
                "status": "running",
                "progress": 50,
                "total": len(parsed_coa),
            }
            gl_result = import_gl_accounts(db, tenant_id, parsed_coa, gl_options)
            run.gl_accounts_imported = gl_result["imported"]
            run.gl_accounts_skipped = gl_result["skipped"]
            if gl_result["errors"]:
                all_errors.extend(gl_result["errors"])
            db.commit()
            yield {
                "step": "gl_accounts",
                "status": "complete",
                "imported": gl_result["imported"],
                "skipped": gl_result["skipped"],
                "errors": gl_result["errors"],
            }
        except Exception as e:
            all_errors.append(f"GL accounts step failed: {str(e)}")
            yield {"step": "gl_accounts", "status": "error", "message": str(e)}
    else:
        yield {"step": "gl_accounts", "status": "skipped", "message": "No COA file provided"}

    # ---- Step 2: Customers + AR Aging ----
    ar_invoices_parsed: list[dict] = []
    ar_customers_parsed: list[dict] = []

    if files.get("ar_aging"):
        try:
            ar_invoices_parsed, ar_customers_parsed = parse_sage_ar_aging(files["ar_aging"])
        except Exception as e:
            all_errors.append(f"AR aging parse failed: {str(e)}")
            yield {"step": "ar_parse", "status": "error", "message": str(e)}

    if files.get("customers"):
        yield {"step": "customers", "status": "running", "progress": 0}
        try:
            parsed_customers = parse_sage_customers(files["customers"])
            yield {
                "step": "customers",
                "status": "running",
                "progress": 50,
                "total": len(parsed_customers),
            }
            cust_result, customer_id_map = import_customers(
                db, tenant_id, parsed_customers, ar_customers_parsed, customer_options
            )
            run.customers_imported = cust_result["imported"]
            run.customers_skipped = cust_result["skipped"]
            if cust_result["errors"]:
                all_errors.extend(cust_result["errors"])
            db.commit()
            yield {
                "step": "customers",
                "status": "complete",
                "imported": cust_result["imported"],
                "skipped": cust_result["skipped"],
                "errors": cust_result["errors"],
            }
        except Exception as e:
            all_errors.append(f"Customers step failed: {str(e)}")
            yield {"step": "customers", "status": "error", "message": str(e)}
    else:
        yield {"step": "customers", "status": "skipped", "message": "No customers file provided"}

    # ---- Step 3: Open AR Invoices ----
    if ar_invoices_parsed and customer_id_map:
        yield {
            "step": "ar_invoices",
            "status": "running",
            "progress": 0,
            "total": len(ar_invoices_parsed),
        }
        try:
            inv_result = import_open_invoices(
                db, tenant_id, ar_invoices_parsed, customer_id_map, cutover_date
            )
            run.ar_invoices_imported = inv_result["imported"]
            run.ar_invoices_skipped = inv_result["skipped"]
            if inv_result.get("warnings"):
                all_warnings.extend(inv_result["warnings"])
            if inv_result.get("errors"):
                all_errors.extend(inv_result["errors"])

            # Calculate total AR balance
            total_ar = sum(
                abs(inv["balance"])
                for inv in ar_invoices_parsed
                if inv["balance"] != Decimal("0")
            )
            run.total_ar_balance = total_ar
            db.commit()
            yield {
                "step": "ar_invoices",
                "status": "complete",
                "imported": inv_result["imported"],
                "skipped": inv_result["skipped"],
                "errors": inv_result.get("errors", []),
                "warnings": inv_result.get("warnings", []),
                "total_ar_balance": float(total_ar),
            }
        except Exception as e:
            all_errors.append(f"AR invoices step failed: {str(e)}")
            yield {"step": "ar_invoices", "status": "error", "message": str(e)}
    elif files.get("ar_aging"):
        yield {
            "step": "ar_invoices",
            "status": "skipped",
            "message": "No customer ID map — run customers import first",
        }
    else:
        yield {"step": "ar_invoices", "status": "skipped", "message": "No AR aging file provided"}

    # ---- Step 4: Vendors ----
    if files.get("vendors"):
        yield {"step": "vendors", "status": "running", "progress": 0}
        try:
            parsed_vendors = parse_sage_vendors(files["vendors"])
            yield {
                "step": "vendors",
                "status": "running",
                "progress": 50,
                "total": len(parsed_vendors),
            }
            vend_result, vendor_id_map = import_vendors(
                db, tenant_id, parsed_vendors, vendor_options
            )
            run.vendors_imported = vend_result["imported"]
            run.vendors_skipped = vend_result["skipped"]
            if vend_result["errors"]:
                all_errors.extend(vend_result["errors"])
            db.commit()
            yield {
                "step": "vendors",
                "status": "complete",
                "imported": vend_result["imported"],
                "skipped": vend_result["skipped"],
                "errors": vend_result["errors"],
            }
        except Exception as e:
            all_errors.append(f"Vendors step failed: {str(e)}")
            yield {"step": "vendors", "status": "error", "message": str(e)}
    else:
        yield {"step": "vendors", "status": "skipped", "message": "No vendors file provided"}

    # ---- Step 5: Open AP Bills ----
    ap_bills_parsed: list[dict] = []
    if files.get("ap_aging"):
        try:
            ap_bills_parsed = parse_sage_ap_aging(files["ap_aging"])
        except Exception as e:
            all_errors.append(f"AP aging parse failed: {str(e)}")
            yield {"step": "ap_parse", "status": "error", "message": str(e)}

    if ap_bills_parsed and vendor_id_map:
        yield {
            "step": "ap_bills",
            "status": "running",
            "progress": 0,
            "total": len(ap_bills_parsed),
        }
        try:
            bill_result = import_open_bills(
                db, tenant_id, ap_bills_parsed, vendor_id_map, cutover_date
            )
            run.ap_bills_imported = bill_result["imported"]
            run.ap_bills_skipped = bill_result["skipped"]
            if bill_result.get("warnings"):
                all_warnings.extend(bill_result["warnings"])
            if bill_result.get("errors"):
                all_errors.extend(bill_result["errors"])

            total_ap = sum(
                abs(b["invoice_balance"])
                for b in ap_bills_parsed
                if b["invoice_balance"] != Decimal("0")
            )
            run.total_ap_balance = total_ap
            db.commit()
            yield {
                "step": "ap_bills",
                "status": "complete",
                "imported": bill_result["imported"],
                "skipped": bill_result["skipped"],
                "errors": bill_result.get("errors", []),
                "warnings": bill_result.get("warnings", []),
                "total_ap_balance": float(total_ap),
            }
        except Exception as e:
            all_errors.append(f"AP bills step failed: {str(e)}")
            yield {"step": "ap_bills", "status": "error", "message": str(e)}
    elif files.get("ap_aging"):
        yield {
            "step": "ap_bills",
            "status": "skipped",
            "message": "No vendor ID map — run vendors import first",
        }
    else:
        yield {"step": "ap_bills", "status": "skipped", "message": "No AP aging file provided"}

    # ---- Apply extension decisions + visibility flags ----
    if extension_decisions:
        from app.services import extension_service as ext_svc
        ext_map = {
            "enable_wastewater": "wastewater",
            "enable_redi_rock": "redi_rock",
            "enable_rosetta": "rosetta",
        }
        for opt_key, ext_key in ext_map.items():
            if extension_decisions.get(opt_key):
                try:
                    ext_svc.enable_extension(db, tenant_id, ext_key)
                except Exception:
                    pass  # Extension may not exist in this env — non-fatal

    # Apply visibility flags to all sage-migrated records
    try:
        _apply_visibility_flags(db, tenant_id)
        db.commit()
    except Exception as vf_err:
        # Non-fatal — visibility can be corrected later
        all_warnings.append(f"Visibility flags not applied: {vf_err}")

    # Compute hidden counts for summary
    _hidden_contractors = 0
    _hidden_products = 0
    try:
        from app.models.customer import Customer as _Cust
        from app.models.product import Product as _Prod
        _hidden_contractors = db.query(_Cust).filter(
            _Cust.company_id == tenant_id,
            _Cust.is_extension_hidden.is_(True),
            _Cust.customer_type == "contractor",
        ).count()
        _hidden_products = db.query(_Prod).filter(
            _Prod.company_id == tenant_id,
            _Prod.is_extension_hidden.is_(True),
        ).count()
    except Exception:
        pass

    # ---- Finalize ----
    final_status = "complete"
    if all_errors:
        final_status = "partial" if (run.customers_imported + run.ar_invoices_imported + run.vendors_imported + run.ap_bills_imported) > 0 else "failed"

    run.status = final_status
    run.completed_at = datetime.now(timezone.utc)
    run.warnings = all_warnings
    run.errors = all_errors
    db.commit()

    # Mark the onboarding checklist item as complete when migration succeeds
    if final_status in ("complete", "partial"):
        try:
            from app.services.onboarding_service import check_completion
            check_completion(db, tenant_id, "data_migration")
            db.commit()
        except Exception as oc_err:
            print(f"WARNING: Could not mark data_migration checklist item complete — {oc_err}")

    yield {
        "status": "complete",
        "run_id": run.id,
        "final_status": final_status,
        "summary": {
            "gl_accounts_imported": run.gl_accounts_imported,
            "gl_accounts_skipped": run.gl_accounts_skipped,
            "customers_imported": run.customers_imported,
            "customers_skipped": run.customers_skipped,
            "ar_invoices_imported": run.ar_invoices_imported,
            "ar_invoices_skipped": run.ar_invoices_skipped,
            "vendors_imported": run.vendors_imported,
            "vendors_skipped": run.vendors_skipped,
            "ap_bills_imported": run.ap_bills_imported,
            "ap_bills_skipped": run.ap_bills_skipped,
            "total_ar_balance": float(run.total_ar_balance) if run.total_ar_balance else 0,
            "total_ap_balance": float(run.total_ap_balance) if run.total_ap_balance else 0,
            "warning_count": len(all_warnings),
            "error_count": len(all_errors),
            "errors": all_errors[:50],
            "warnings": all_warnings[:50],
            "hidden_contractors": _hidden_contractors,
            "hidden_products": _hidden_products,
        },
    }


# ---------------------------------------------------------------------------
# Status + Rollback
# ---------------------------------------------------------------------------


def get_migration_status(db, tenant_id: str) -> dict | None:
    """Return the latest DataMigrationRun for a tenant, as a dict."""
    run = (
        db.query(DataMigrationRun)
        .filter(DataMigrationRun.tenant_id == tenant_id)
        .order_by(DataMigrationRun.started_at.desc())
        .first()
    )
    if run is None:
        return None

    return {
        "id": run.id,
        "status": run.status,
        "cutover_date": str(run.cutover_date),
        "source_system": run.source_system,
        "gl_accounts_imported": run.gl_accounts_imported,
        "gl_accounts_skipped": run.gl_accounts_skipped,
        "customers_imported": run.customers_imported,
        "customers_skipped": run.customers_skipped,
        "ar_invoices_imported": run.ar_invoices_imported,
        "ar_invoices_skipped": run.ar_invoices_skipped,
        "vendors_imported": run.vendors_imported,
        "vendors_skipped": run.vendors_skipped,
        "ap_bills_imported": run.ap_bills_imported,
        "ap_bills_skipped": run.ap_bills_skipped,
        "total_ar_balance": float(run.total_ar_balance) if run.total_ar_balance else None,
        "total_ap_balance": float(run.total_ap_balance) if run.total_ap_balance else None,
        "warnings": run.warnings or [],
        "errors": run.errors or [],
        "started_at": run.started_at.isoformat() if run.started_at else None,
        "completed_at": run.completed_at.isoformat() if run.completed_at else None,
        "initiated_by": run.initiated_by,
        "rolled_back_at": run.rolled_back_at.isoformat() if run.rolled_back_at else None,
        "rolled_back_by": run.rolled_back_by,
    }


def rollback_migration(
    db,
    tenant_id: str,
    run_id: str,
    rolled_back_by: str,
) -> dict:
    """Roll back a migration run by deleting all sage-migrated records.

    Deletes:
    - Customers with sage_customer_id set (for this tenant)
    - Invoices with sage_invoice_id set (for this tenant)
    - VendorBills with source == "sage_migration" (for this tenant)
    - TenantGLMappings imported in this run (identified by provider_account_id matching)
    """
    rolled_back_records = 0

    # Find the run
    run = db.query(DataMigrationRun).filter(
        DataMigrationRun.id == run_id,
        DataMigrationRun.tenant_id == tenant_id,
    ).first()
    if run is None:
        raise ValueError(f"Migration run {run_id} not found for this tenant")

    # Delete invoice lines for sage invoices first (FK constraint)
    sage_invoices = (
        db.query(Invoice)
        .filter(
            Invoice.company_id == tenant_id,
            Invoice.sage_invoice_id.isnot(None),
        )
        .all()
    )
    for invoice in sage_invoices:
        for line in invoice.lines:
            db.delete(line)
        db.delete(invoice)
        rolled_back_records += 1

    # Delete vendor bill lines for sage bills first
    sage_bills = (
        db.query(VendorBill)
        .filter(
            VendorBill.company_id == tenant_id,
            VendorBill.source == "sage_migration",
        )
        .all()
    )
    for bill in sage_bills:
        for line in bill.lines:
            db.delete(line)
        db.delete(bill)
        rolled_back_records += 1

    # Delete customers with sage_customer_id
    sage_customers = (
        db.query(Customer)
        .filter(
            Customer.company_id == tenant_id,
            Customer.sage_customer_id.isnot(None),
        )
        .all()
    )
    for customer in sage_customers:
        db.delete(customer)
        rolled_back_records += 1

    # Delete vendors with sage_vendor_id
    sage_vendors = (
        db.query(Vendor)
        .filter(
            Vendor.company_id == tenant_id,
            Vendor.sage_vendor_id.isnot(None),
        )
        .all()
    )
    for vendor in sage_vendors:
        db.delete(vendor)
        rolled_back_records += 1

    # Delete GL mappings imported in this run
    # We identify them by provider_account_id being numeric (sage account numbers)
    sage_gl_mappings = (
        db.query(TenantGLMapping)
        .filter(
            TenantGLMapping.tenant_id == tenant_id,
            TenantGLMapping.provider_account_id.isnot(None),
        )
        .all()
    )
    for mapping in sage_gl_mappings:
        # Only delete ones that look like Sage account numbers (numeric)
        pid = mapping.provider_account_id or ""
        if pid.isdigit():
            db.delete(mapping)
            rolled_back_records += 1

    # Update run status
    run.status = "rolled_back"
    run.rolled_back_at = datetime.now(timezone.utc)
    run.rolled_back_by = rolled_back_by

    db.commit()

    return {"rolled_back_records": rolled_back_records, "run_id": run_id}


# ---------------------------------------------------------------------------
# DataMigrationService facade class (optional convenience wrapper)
# ---------------------------------------------------------------------------


# ---------------------------------------------------------------------------
# Extension Content Detection
# ---------------------------------------------------------------------------

_WASTEWATER_TERMS = [
    "septic", "tank", "wastewater", "distribution box", "pump chamber",
    "riser", "filter", "leach", "cesspool", "1000 gal", "1500 gal", "2000 gal",
]
_REDI_ROCK_TERMS = [
    "redi-rock", "redi rock", "retaining", "wall block", "standard block", "large block",
]
_ROSETTA_TERMS = [
    "rosetta", "hardscape", "paver", "decorative concrete",
]


def _name_contains(name: str, terms: list[str]) -> bool:
    low = name.lower()
    return any(t in low for t in terms)


def detect_extension_content(
    parsed_customers: list[dict],
    parsed_products: list[dict],
) -> dict:
    """Scan parsed records and return a summary of extension-requiring content.

    Used by the parse endpoint so the frontend can show Step 2.5.
    """
    contractor_names = [
        c["name"] for c in parsed_customers if c.get("customer_type") == "contractor"
    ]
    wastewater_names = [
        p["name"] for p in parsed_products if _name_contains(p.get("name", ""), _WASTEWATER_TERMS)
    ]
    redi_rock_names = [
        p["name"] for p in parsed_products if _name_contains(p.get("name", ""), _REDI_ROCK_TERMS)
    ]
    rosetta_names = [
        p["name"] for p in parsed_products if _name_contains(p.get("name", ""), _ROSETTA_TERMS)
    ]

    has_content = bool(contractor_names or wastewater_names or redi_rock_names or rosetta_names)

    suggested = []
    if contractor_names or wastewater_names:
        suggested.append("wastewater")
    if contractor_names or redi_rock_names:
        suggested.append("redi_rock")

    return {
        "contractor_customers": {
            "count": len(contractor_names),
            "sample_names": contractor_names[:5],
            "suggested_extensions": suggested if contractor_names else [],
        },
        "wastewater_products": {
            "count": len(wastewater_names),
            "sample_names": wastewater_names[:5],
            "suggested_extension": "wastewater",
        },
        "redi_rock_products": {
            "count": len(redi_rock_names),
            "sample_names": redi_rock_names[:5],
            "suggested_extension": "redi_rock",
        },
        "rosetta_products": {
            "count": len(rosetta_names),
            "sample_names": rosetta_names[:5],
            "suggested_extension": "rosetta",
        },
        "has_extension_content": has_content,
    }


def _apply_visibility_flags(db, tenant_id: str) -> None:
    """After import, apply is_extension_hidden to all imported customers and products.

    Checks which product-line extensions are now active and sets flags accordingly.
    Called at the end of run_full_migration after extension_decisions are applied.
    """
    from app.models.customer import Customer
    from app.models.product import Product
    from app.models.tenant_extension import TenantExtension

    PRODUCT_LINE_KEYS = ("wastewater", "redi_rock", "rosetta")

    active_exts = {
        te.extension_key
        for te in db.query(TenantExtension).filter(
            TenantExtension.tenant_id == tenant_id,
            TenantExtension.extension_key.in_(PRODUCT_LINE_KEYS),
            TenantExtension.status == "active",
        ).all()
    }
    any_product_line_active = bool(active_exts)

    # ── Customers ─────────────────────────────────────────────────────────────
    customers = db.query(Customer).filter(
        Customer.company_id == tenant_id,
        Customer.source == "sage_migration",
    ).all()

    for c in customers:
        if c.customer_type == "funeral_home":
            c.is_extension_hidden = False
            c.visibility_requires_extension = None
        elif c.customer_type == "contractor":
            c.visibility_requires_extension = "any_product_line"
            c.is_extension_hidden = not any_product_line_active
        # other types: leave as-is (default false)

    # ── Products ──────────────────────────────────────────────────────────────
    products = db.query(Product).filter(
        Product.company_id == tenant_id,
        Product.source == "sage_migration",
    ).all()

    for p in products:
        name = p.name or ""
        if _name_contains(name, _WASTEWATER_TERMS):
            p.visibility_requires_extension = "wastewater"
            p.is_extension_hidden = "wastewater" not in active_exts
        elif _name_contains(name, _REDI_ROCK_TERMS):
            p.visibility_requires_extension = "redi_rock"
            p.is_extension_hidden = "redi_rock" not in active_exts
        elif _name_contains(name, _ROSETTA_TERMS):
            p.visibility_requires_extension = "rosetta"
            p.is_extension_hidden = "rosetta" not in active_exts
        else:
            p.visibility_requires_extension = None
            p.is_extension_hidden = False

    db.flush()


class DataMigrationService:
    """Thin facade so route code can call DataMigrationService.parse_coa() etc."""

    @staticmethod
    def parse_coa(content: bytes) -> list[dict]:
        return parse_sage_coa(content)

    @staticmethod
    def parse_customers(content: bytes) -> list[dict]:
        return parse_sage_customers(content)

    @staticmethod
    def parse_vendors(content: bytes) -> list[dict]:
        return parse_sage_vendors(content)

    @staticmethod
    def parse_ar_aging(content: bytes) -> tuple[list[dict], list[dict]]:
        return parse_sage_ar_aging(content)

    @staticmethod
    def parse_ap_aging(content: bytes) -> list[dict]:
        return parse_sage_ap_aging(content)

    @staticmethod
    def get_migration_status(db, tenant_id: str) -> dict | None:
        return get_migration_status(db, tenant_id)

    @staticmethod
    def rollback_migration(db, tenant_id: str, run_id: str, rolled_back_by: str) -> dict:
        return rollback_migration(db, tenant_id, run_id, rolled_back_by)

    @staticmethod
    def detect_extension_content(
        parsed_customers: list[dict],
        parsed_products: list[dict],
    ) -> dict:
        return detect_extension_content(parsed_customers, parsed_products)

    @staticmethod
    def run_full_migration(
        db,
        tenant_id: str,
        files: dict,
        options: dict,
        cutover_date: date,
        initiated_by: str = "owner",
        extension_decisions: dict | None = None,
    ) -> Generator[dict, None, None]:
        return run_full_migration(
            db, tenant_id, files, options, cutover_date, initiated_by, extension_decisions
        )
